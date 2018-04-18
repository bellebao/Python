# -*- coding: utf-8 -*-
"""
Created on Tue Mar 20 16:07:45 2018

Note: Calculate a cycle time by the PAR status. The end time is inwork/close,
the initial time is raised. CT is the time buffer between them.

@author: Belle Bao
"""

import sys
import os
import string
#import win32com.client
import re
import datetime
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color

f_output = r"C:\IOSystems\A300\PAR\PAR_Analysis\CycleTime.txt"
f_result_file = r"C:\IOSystems\A300\PAR\PAR_Analysis\Par\CycleTime.xlsx"  # output file for CT
path = "C:\IOSystems\A300\PAR\PAR_Analysis\Par\PAR"   # input file where have FCR
#--------------------------------CONSTANTS-------------------------------------
wdFormatText = 2

#--------------------------------GLOBALS---------------------------------------
TimeRaisedTemp = ''
TimeInworkTemp = ''
PAR_NoTemp= ''
CTtemp = ''
templine = []
InreviewTime = ''
CycleTime = {}
filenum = 0
BOETime = {}
ListBOEKey = []

def ParseFile(path):
    
    if os.path.isdir(path):
        filenum = len(os.listdir(path))
        for sub_dir in os.listdir(path):
            filename = os.path.join(path, sub_dir)
            ParseFile(filename)
    else:
        TraceParser(path)
    

def TraceParser(filename):
    
    global filenum
    #global CycleTime
    fileext = ""
    actionline = []
    actionflag = 0
    raisedTimeBeforFormat = ""
    InreviewTimeBeforFormat = ""
    CloseTimeBeforFormat = ""
    raisedTimeFormat = ""
    InreviewTimeFormat = ""
    CloseTimeFormat = ""
    CTReview = ""
    CTClose = ""
    

    #filename = str.lower(filename)
    index = filename.rfind('.')
    
    if (index != -1):
        fileext = filename[index:]
    
    if (fileext in [".par", ".fcr", ".scn",".txt",".TXT"]):
        
        filenum = filenum + 1    
        
        # get the filename without the directory
        abbrname = os.path.split(filename)[1]
        PAR_NoTemp = abbrname.split('.')[0]
        
        # start parsing the file
        parsefile = open(filename, 'r')
        keyword = "ACTION HISTORY"
        count = 0
        for (line) in parsefile.readlines():
            count = count +1
            if (keyword in line):
                actionflag = 1
            
            if(actionflag == 1):
                actionline.append(line)
                #print (actionline)
        if(actionflag == 1):        
            raisedTime = re.search(r"(\d{2}-(DEC|NOV|OCT|SEP|AUG|JUL|JUN|MAY|APR|MAR|FEB|JAN)-\d{4})", actionline[1]) 
            raisedTimeBeforFormat = raisedTime.group(0)
            raisedTimeFormat = transferdate(raisedTimeBeforFormat)
        #print (raisedTime.group(0))
        count = 0
        for line in actionline:     
            count = count +1
            if ("Actioned document from ANALYZE EFFORT to IN WORK" in line):
                #continue
                InreviewTime = re.search(r"(\d{2}-(DEC|NOV|OCT|SEP|AUG|JUL|JUN|MAY|APR|MAR|FEB|JAN)-\d{4})", actionline[count-3])
                #print (InreviewTime)
                InreviewTimeBeforFormat = InreviewTime.group(0)
                InreviewTimeFormat = transferdate(InreviewTimeBeforFormat)

            if("Actioned document from IN WORK to CLOSED" in line): 
                #continue
                CloseTime = re.search(r"(\d{2}-(DEC|NOV|OCT|SEP|AUG|JUL|JUN|MAY|APR|MAR|FEB|JAN)-\d{4})", actionline[count-3])
                #print (InreviewTime)
                CloseTimeBeforFormat = CloseTime.group(0)
                CloseTimeFormat = transferdate(CloseTimeBeforFormat)
           
            if (InreviewTimeBeforFormat != "" ):
                CTReview = calculateCT(raisedTimeFormat, InreviewTimeFormat)
            if (CloseTimeBeforFormat!= "" ):
                CTClose = calculateCT(raisedTimeFormat, CloseTimeFormat)
                
        CycleTime.update({PAR_NoTemp:[raisedTimeBeforFormat,InreviewTimeBeforFormat,CloseTimeBeforFormat, CTReview, CTClose, raisedTimeFormat, InreviewTimeFormat, CloseTimeFormat]})


def calculateCT(raisedTimeFormat, InreviewTimeFormat):
    date1 =  raisedTimeFormat.split('-') 
    date2 =  InreviewTimeFormat.split('-') 
    d1 = datetime.datetime(int(date1[2]),int(date1[1]),int(date1[0]))
    d2 = datetime.datetime(int(date2[2]),int(date2[1]),int(date2[0]))
    return (d2-d1).days    

def transferdate(dateinStr):
    if "JAN" in dateinStr:
        dateinStr=dateinStr.replace('JAN','01')
        return dateinStr
    if "FEB" in dateinStr:
        dateinStr=dateinStr.replace('FEB','02')        
        return dateinStr
    if "MAR" in dateinStr:
        dateinStr=dateinStr.replace('MAR','03')        
        return dateinStr
    if "APR" in dateinStr:
        dateinStr=dateinStr.replace('APR','04')
        return dateinStr
    if "MAY" in dateinStr:
        dateinStr=dateinStr.replace('MAY','05')        
        return dateinStr
    if "JUN" in dateinStr:
        dateinStr=dateinStr.replace('JUN','06')        
        return dateinStr
    if "JUL" in dateinStr:
        dateinStr=dateinStr.replace('JUL','07')
        return dateinStr
    if "AUG" in dateinStr:
        dateinStr=dateinStr.replace('AUG','08')        
        return dateinStr
    if "SEP" in dateinStr:
        dateinStr=dateinStr.replace('SEP','09')        
        return dateinStr
    if "OCT" in dateinStr:
        dateinStr=dateinStr.replace('OCT','10')
        return dateinStr
    if "NOV" in dateinStr:
        dateinStr=dateinStr.replace('NOV','11')        
        return dateinStr
    if "DEC" in dateinStr:
        dateinStr=dateinStr.replace('DEC','12')        
        return dateinStr    
def cli():  
    global filenum

    ParseFile(path)
    #print (CycleTime)
    #ff = open(f_output, 'w')
    #ff.writelines( "%s\t%s\t%s\t%s\n" % ("PAR","RAISED Time","Review Time","Cycle Time") )
    #for key in CycleTime:
    #    ff.writelines( "%s\t%s\t%s\t%s\n" % (key, CycleTime[key][0], CycleTime[key][1], CycleTime[key][2]) )
    #ff.close()
    #from openpyxl import Workbook
    BOEFORALL= 0 
    REVIEWNo = 0
    
    wb = Workbook()
    ###############################################
    ws1 = wb.active
    ws1.title = "CycleTime_AllPAR"
    
    wsheader = ["PAR No.", "RAISED Time","Review Time","Close Time", "Cycle Time for review(Day)", "Cycle Time for closure(Day)"]
    ws1.append(wsheader)
    for key in CycleTime:
        row = [key]
        row.append(CycleTime[key][0])
        row.append(CycleTime[key][1])
        row.append(CycleTime[key][2])
        row.append(CycleTime[key][3])
        row.append(CycleTime[key][4])
        ws1.append(row)
        
        if CycleTime[key][1] != "":
            REVIEWNo = REVIEWNo + 1
            BOEFORALL = BOEFORALL + CycleTime[key][3]
            # CALCULATE BOE FOR EACH MONTH
            BOEKEY = CycleTime[key][1].split('-')[1] + CycleTime[key][1].split('-')[2]
            if BOEKEY in ListBOEKey:
                CycleTimeTemp = CycleTime[key][3]+BOETime.get(BOEKEY)[0]
                CRNoTemp = 1 + BOETime.get(BOEKEY)[1]
                BOEaverage = CycleTimeTemp/CRNoTemp
                BOETime.update( {BOEKEY:[CycleTimeTemp,CRNoTemp,BOEaverage] })
            else:
                ListBOEKey.append(BOEKEY)
            if BOEKEY not in BOETime:
                BOETime.update( {BOEKEY:[CycleTime[key][3],1,CycleTime[key][3]] })
        
        # CALCULATE BOE FOR EACH MONTH
    if REVIEWNo != 0:
        BOEAverage = BOEFORALL/REVIEWNo    
    
    ws2 = wb.create_sheet()
    ws2.title = "CycleTime_BOE"
    
    wsheader = ["Baseline", "Review BOE for A PAR(Day)", "PAR No.", "Review time in Total(Day)"]
    ws2.append(wsheader)
    row = ["2017 - 2018 Year"]
    row.append(BOEAverage)
    row.append(REVIEWNo)
    row.append(BOEFORALL)
    ws2.append(row)
    for key in BOETime:
        row = [key]
        row.append(BOETime[key][2])
        row.append(BOETime[key][1])
        row.append(BOETime[key][0])
        ws2.append(row)
    print ("-->Saving File: "+f_result_file)
    wb.save(f_result_file)          
    print ("-->Done!")

if __name__ == "__main__":
    cli()        
        
        
        
        